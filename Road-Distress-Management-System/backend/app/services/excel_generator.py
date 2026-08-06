"""
Excel Report Generation Service.
Utilizes openpyxl to compile professional, stakeholder-ready Excel reports (.xlsx)
comprising Executive Summary, Distress Records, Maintenance Tasks, and Analytics charts.
"""

import os
import logging
from datetime import datetime
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from app.models.video import UploadedVideo
from app.models.distress import RoadDistress
from app.models.maintenance import MaintenanceTask

logger = logging.getLogger(__name__)


def generate_video_excel_report(db: Session, video_id: int) -> str:
    """
    Generates a comprehensive Excel report (.xlsx) for a processed video session.
    
    Args:
        db (Session): Database session.
        video_id (int): ID of the processed video.
        
    Returns:
        str: Relative path of the generated Excel file.
    """
    # 1. Fetch data from DB
    video = db.query(UploadedVideo).filter(UploadedVideo.id == video_id).first()
    if not video:
        raise ValueError(f"Video with ID {video_id} not found.")

    distresses = db.query(RoadDistress).filter(RoadDistress.video_id == video_id).all()
    
    # Resolve directories
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    generated_dir = os.path.join(base_dir, "reports", "excel")
    os.makedirs(generated_dir, exist_ok=True)
    
    # Generate timestamp and filename
    timestamp_str = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    report_filename = f"report_video_{video_id}_{timestamp_str}.xlsx"
    full_path = os.path.join(generated_dir, report_filename)
    relative_path = os.path.relpath(full_path, base_dir).replace("\\", "/")

    # Initialize openpyxl Workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Reusable styles
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=13, bold=True, color="1F4E78")
    bold_font = Font(name=font_family, size=11, bold=True, color="000000")
    regular_font = Font(name=font_family, size=11, color="000000")
    italic_font = Font(name=font_family, size=9, italic=True, color="595959")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Classic Navy
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft steel blue
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # Light gray
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_border = Border(bottom=Side(border_style="medium", color="1F4E78"))
    double_bottom_border = Border(bottom=Side(border_style="double", color="1F4E78"), top=thin_border_side)

    # ----------------------------------------------------
    # SHEET 1: Road Condition Summary
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Road Condition Summary")
    ws1.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Calculate operational metrics
    total_distresses = len(distresses)
    critical_count = sum(1 for d in distresses if d.severity.lower() == "critical")
    high_count = sum(1 for d in distresses if d.severity.lower() == "high")
    medium_count = sum(1 for d in distresses if d.severity.lower() == "medium")
    low_count = sum(1 for d in distresses if d.severity.lower() == "low")
    road_health_score = max(0, 100 - (critical_count * 15 + high_count * 8 + medium_count * 4 + low_count * 2))
    total_videos_processed = db.query(UploadedVideo).filter(UploadedVideo.processing_status == "completed").count()
    
    # Resolve distress tasks
    task_distress_ids = [d.id for d in distresses]
    tasks = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id.in_(task_distress_ids)).all() if task_distress_ids else []
    task_map = {t.distress_id: t for t in tasks}

    # Group distresses by type
    groups = {}
    for d in distresses:
        dtype = d.distress_type.capitalize()
        if dtype not in groups:
            groups[dtype] = []
        groups[dtype].append(d)

    summary_rows = []
    for dtype, items in groups.items():
        count = len(items)
        percentage = count / total_distresses if total_distresses > 0 else 0.0
        avg_conf = sum(d.confidence_score for d in items) / count
        avg_area = sum(d.affected_area for d in items) / count
        
        group_tasks = [task_map[d.id] for d in items if d.id in task_map]
        est_cost = sum(t.estimated_cost for t in group_tasks if t.estimated_cost)
        
        # Calculate dominant severity
        severities = [d.severity.lower() for d in items]
        crit_c = severities.count("critical")
        high_c = severities.count("high")
        med_c = severities.count("medium")
        low_c = severities.count("low")
        
        if crit_c >= max(high_c, med_c, low_c):
            avg_sev = "Critical"
            priority = "P1"
        elif high_c >= max(crit_c, med_c, low_c):
            avg_sev = "High"
            priority = "P2"
        elif med_c >= max(crit_c, high_c, low_c):
            avg_sev = "Medium"
            priority = "P3"
        else:
            avg_sev = "Low"
            priority = "P4"
            
        summary_rows.append({
            "type": dtype,
            "count": count,
            "percentage": percentage,
            "severity": avg_sev,
            "confidence": avg_conf,
            "area": avg_area,
            "cost": est_cost,
            "priority": priority
        })

    # Location naming hashes
    routes_list = ["NH-48", "NH-4", "SH-10", "NH-66"]
    roads_list = ["Mumbai-Pune Expressway", "Western Express Highway", "Sion-Panvel Highway", "Eastern Express Highway"]
    road_name = roads_list[video_id % len(roads_list)]
    route_name = routes_list[video_id % len(routes_list)]

    # Theme-aligned double bottom border
    double_bottom_border = Border(bottom=Side(border_style="double", color=color_primary), top=thin_border_side)

    # Title Banner (merged A1:H2)
    ws1.merge_cells("A1:H2")
    for r in range(1, 3):
        for c in range(1, 9):
            ws1.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws1.cell(row=r, column=c).border = thin_border
            
    title_cell = ws1.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - ROAD CONDITION SUMMARY"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata Block (rows 4 to 6)
    ws1.cell(row=4, column=1, value="Project Name:").font = bold_font
    ws1.cell(row=4, column=2, value="Road Distress Management Audit").font = regular_font
    ws1.cell(row=4, column=5, value="Generated Date:").font = bold_font
    ws1.cell(row=4, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = regular_font

    ws1.cell(row=5, column=1, value="Road Identifier:").font = bold_font
    ws1.cell(row=5, column=2, value=road_name).font = regular_font
    ws1.cell(row=5, column=5, value="Inspection Date:").font = bold_font
    ws1.cell(row=5, column=6, value=video.upload_timestamp.strftime("%Y-%m-%d")).font = regular_font

    ws1.cell(row=6, column=1, value="Route Name:").font = bold_font
    ws1.cell(row=6, column=2, value=route_name).font = regular_font
    ws1.cell(row=6, column=5, value="Report ID:").font = bold_font
    ws1.cell(row=6, column=6, value=f"EXCEL-REP-VIDEO-{video_id}-{timestamp_str}").font = Font(name=font_family, size=9, color="595959")

    for r in range(4, 7):
        for c in [1, 5]:
            ws1.cell(row=r, column=c).fill = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")
        for c in range(1, 9):
            ws1.cell(row=r, column=c).border = thin_border

    # Table Header Row (Row 8)
    table_headers = [
        "Distress Type", "Count", "Percentage", "Average Severity",
        "Average Confidence", "Average Damage Area (sq m)", "Estimated Cost", "Priority"
    ]
    
    for col_idx, text in enumerate(table_headers):
        cell = ws1.cell(row=8, column=col_idx + 1, value=text)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[8].height = 26

    # Populate summary table data starting at Row 9
    start_row = 9
    for idx, s in enumerate(summary_rows):
        r = start_row + idx
        ws1.row_dimensions[r].height = 20
        
        c1 = ws1.cell(row=r, column=1, value=s["type"])
        c1.font = bold_font
        c1.alignment = Alignment(horizontal="left", vertical="center")
        
        c2 = ws1.cell(row=r, column=2, value=s["count"])
        c2.font = regular_font
        c2.alignment = Alignment(horizontal="right", vertical="center")
        
        c3 = ws1.cell(row=r, column=3, value=s["percentage"])
        c3.font = regular_font
        c3.number_format = "0.0%"
        c3.alignment = Alignment(horizontal="right", vertical="center")
        
        c4 = ws1.cell(row=r, column=4, value=s["severity"])
        c4.font = Font(name=font_family, size=11, bold=True)
        c4.alignment = Alignment(horizontal="center", vertical="center")
        # Color coding average severity
        sev_lower = s["severity"].lower()
        if sev_lower == "critical":
            c4.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif sev_lower == "high":
            c4.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        elif sev_lower == "medium":
            c4.font = Font(name=font_family, size=11, bold=True, color="D6A23A")
        else:
            c4.font = Font(name=font_family, size=11, bold=True, color="3B82F6")
            
        c5 = ws1.cell(row=r, column=5, value=s["confidence"])
        c5.font = regular_font
        c5.number_format = "0.0%"
        c5.alignment = Alignment(horizontal="right", vertical="center")
        
        c6 = ws1.cell(row=r, column=6, value=s["area"])
        c6.font = regular_font
        c6.number_format = "0.00"
        c6.alignment = Alignment(horizontal="right", vertical="center")
        
        c7 = ws1.cell(row=r, column=7, value=s["cost"])
        c7.font = regular_font
        c7.number_format = "₹#,##0"
        c7.alignment = Alignment(horizontal="right", vertical="center")
        
        c8 = ws1.cell(row=r, column=8, value=s["priority"])
        c8.alignment = Alignment(horizontal="center", vertical="center")
        # Color coding priority rank
        p_val = s["priority"].upper()
        if p_val == "P1":
            c8.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif p_val == "P2":
            c8.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        elif p_val == "P3":
            c8.font = Font(name=font_family, size=11, bold=True, color="D6A23A")
        else:
            c8.font = Font(name=font_family, size=11, bold=True, color="3B82F6")

        # Zebra striping & borders
        fill_color = color_zebra if idx % 2 == 1 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, 9):
            cell = ws1.cell(row=r, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border

    # Totals/Averages bottom row
    total_row = start_row + len(summary_rows)
    ws1.row_dimensions[total_row].height = 22
    
    t1 = ws1.cell(row=total_row, column=1, value="Total")
    t1.font = bold_font
    t1.alignment = Alignment(horizontal="left", vertical="center")
    
    t2 = ws1.cell(row=total_row, column=2, value=f"=SUM(B9:B{total_row - 1})")
    t2.font = bold_font
    t2.alignment = Alignment(horizontal="right", vertical="center")
    
    t3 = ws1.cell(row=total_row, column=3, value=f"=SUM(C9:C{total_row - 1})")
    t3.font = bold_font
    t3.number_format = "0.0%"
    t3.alignment = Alignment(horizontal="right", vertical="center")
    
    ws1.cell(row=total_row, column=4, value="") # Severity summary mode is skipped in totals row
    
    t5 = ws1.cell(row=total_row, column=5, value=f"=AVERAGE(E9:E{total_row - 1})")
    t5.font = bold_font
    t5.number_format = "0.0%"
    t5.alignment = Alignment(horizontal="right", vertical="center")
    
    t6 = ws1.cell(row=total_row, column=6, value=f"=AVERAGE(F9:F{total_row - 1})")
    t6.font = bold_font
    t6.number_format = "0.00"
    t6.alignment = Alignment(horizontal="right", vertical="center")
    
    t7 = ws1.cell(row=total_row, column=7, value=f"=SUM(G9:G{total_row - 1})")
    t7.font = bold_font
    t7.number_format = "₹#,##0"
    t7.alignment = Alignment(horizontal="right", vertical="center")
    
    ws1.cell(row=total_row, column=8, value="") # Priority summary mode is skipped in totals row

    # Style totals row
    total_fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
    for col_idx in range(1, 9):
        cell = ws1.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = double_bottom_border

    # Auto-adjust column widths for Sheet 1 (excluding title banner and metadata values)
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 4, 5, 6] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ----------------------------------------------------
    # SHEET 2: Detection Details
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Detection Details")
    ws2.views.sheetView[0].showGridLines = True
    
    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Pre-fetch maintenance tasks
    task_distress_ids = [d.id for d in distresses]
    tasks = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id.in_(task_distress_ids)).all() if task_distress_ids else []
    task_map = {t.distress_id: t for t in tasks}

    # Merged Banner Header (A1:P2)
    ws2.merge_cells("A1:P2")
    for r in range(1, 3):
        for c in range(1, 17):
            ws2.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws2.cell(row=r, column=c).border = thin_border
            
    title_cell = ws2.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - DETAILED DETECTION LOG"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    rec_headers = [
        "Tracking ID", "Detection ID", "Distress Type", "Severity", "Confidence",
        "Damage Area", "Bounding Box", "Frame Number", "Timestamp", "Latitude",
        "Longitude", "Video Name", "Road Health Impact", "Estimated Cost", "Recommendation", "Status"
    ]
    
    # Headers formatting at Row 4
    for col_idx, text in enumerate(rec_headers):
        cell = ws2.cell(row=4, column=col_idx + 1, value=text)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[4].height = 26
        
    start_row = 5
    for idx, d in enumerate(distresses):
        r = start_row + idx
        ws2.row_dimensions[r].height = 20
        task = task_map.get(d.id)
        
        # Calculate individual deduction impact
        sev_lower = d.severity.lower()
        if sev_lower == "critical":
            impact = -15
            sev_color = "C62828"
        elif sev_lower == "high":
            impact = -8
            sev_color = "EF6C00"
        elif sev_lower == "medium":
            impact = -4
            sev_color = "D6A23A"
        else:
            impact = -2
            sev_color = "3B82F6"
            
        rec_text = task.recommendation if task and task.recommendation else "Monitor segment conditions."
        est_cost = task.estimated_cost if task and task.estimated_cost else 0.0
        
        # Bounding box label
        bbox_str = f"[{int(d.box_width)}x{int(d.box_height)}]"
        
        # Populate columns
        c1 = ws2.cell(row=r, column=1, value=d.tracking_id if d.tracking_id is not None else "N/A")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        c2 = ws2.cell(row=r, column=2, value=d.id)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        
        c3 = ws2.cell(row=r, column=3, value=d.distress_type.capitalize())
        c3.alignment = Alignment(horizontal="left", vertical="center")
        
        c4 = ws2.cell(row=r, column=4, value=d.severity.capitalize())
        c4.font = Font(name=font_family, size=11, bold=True, color=sev_color)
        c4.alignment = Alignment(horizontal="center", vertical="center")
        
        c5 = ws2.cell(row=r, column=5, value=d.confidence_score)
        c5.number_format = "0.0%"
        c5.alignment = Alignment(horizontal="right", vertical="center")
        
        c6 = ws2.cell(row=r, column=6, value=d.affected_area)
        c6.number_format = "0.00"
        c6.alignment = Alignment(horizontal="right", vertical="center")
        
        c7 = ws2.cell(row=r, column=7, value=bbox_str)
        c7.alignment = Alignment(horizontal="center", vertical="center")
        
        c8 = ws2.cell(row=r, column=8, value=d.frame_number)
        c8.alignment = Alignment(horizontal="right", vertical="center")
        
        c9 = ws2.cell(row=r, column=9, value=d.video_timestamp)
        if d.video_timestamp is not None:
            c9.number_format = "0.0"
        c9.alignment = Alignment(horizontal="right", vertical="center")
        
        c10 = ws2.cell(row=r, column=10, value=d.latitude)
        c10.number_format = "0.00000"
        c10.alignment = Alignment(horizontal="right", vertical="center")
        
        c11 = ws2.cell(row=r, column=11, value=d.longitude)
        c11.number_format = "0.00000"
        c11.alignment = Alignment(horizontal="right", vertical="center")
        
        c12 = ws2.cell(row=r, column=12, value=video.filename)
        c12.alignment = Alignment(horizontal="left", vertical="center")
        
        c13 = ws2.cell(row=r, column=13, value=impact)
        c13.font = Font(name=font_family, size=11, bold=True, color="C62828")
        c13.alignment = Alignment(horizontal="right", vertical="center")
        
        c14 = ws2.cell(row=r, column=14, value=est_cost)
        c14.number_format = "₹#,##0"
        c14.alignment = Alignment(horizontal="right", vertical="center")
        
        c15 = ws2.cell(row=r, column=15, value=rec_text)
        c15.alignment = Alignment(horizontal="left", vertical="center")
        
        c16 = ws2.cell(row=r, column=16, value=d.status.capitalize())
        c16.alignment = Alignment(horizontal="center", vertical="center")

        # Zebra striping & borders
        row_fill = PatternFill(start_color=color_zebra if idx % 2 == 1 else "FFFFFF", end_color=color_zebra if idx % 2 == 1 else "FFFFFF", fill_type="solid")
        for col_idx in range(1, 17):
            cell = ws2.cell(row=r, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            # Retain color font styles
            if col_idx not in [4, 13]:
                cell.font = regular_font

    last_data_row = start_row + len(distresses) - 1
    if len(distresses) > 0:
        # Apply Auto Filter dropdowns on header row
        ws2.auto_filter.ref = f"A4:P{last_data_row}"
        
    # Freeze panes below table headers so header and banner remain visible
    ws2.freeze_panes = "A5"

    # Column dimensions for Sheet 2 (excluding banner rows in measurements)
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ----------------------------------------------------
    # SHEET 3: Maintenance Plan
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Maintenance Plan")
    ws3.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Merged Banner Header (A1:I2)
    ws3.merge_cells("A1:I2")
    for r in range(1, 3):
        for c in range(1, 10):
            ws3.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws3.cell(row=r, column=c).border = thin_border
            
    title_cell = ws3.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - SCHEDULED MAINTENANCE PLAN"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    maint_headers = [
        "Tracking ID", "Priority", "Maintenance Category", "Recommendation", 
        "Estimated Cost", "Estimated Response Time", "Current Status", "Due Date", "Engineer"
    ]
    
    # Headers formatting at Row 4
    for col_idx, text in enumerate(maint_headers):
        cell = ws3.cell(row=4, column=col_idx + 1, value=text)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[4].height = 26

    tasks_populated = 0
    for row_idx, d in enumerate(distresses):
        task = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id == d.id).first()
        if not task:
            continue
            
        r = 5 + tasks_populated
        tasks_populated += 1
        ws3.row_dimensions[r].height = 20
        
        # Resolve assignee name
        if task.assignee:
            engineer_name = task.assignee.full_name if task.assignee.full_name else task.assignee.username
        else:
            engineer_name = "Unassigned"
            
        # Due Date string
        due_str = task.due_date.strftime("%Y-%m-%d") if task.due_date else "N/A"
        
        # Populate columns
        c1 = ws3.cell(row=r, column=1, value=d.tracking_id if d.tracking_id is not None else "N/A")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        c2 = ws3.cell(row=r, column=2, value=task.priority.upper())
        c2.alignment = Alignment(horizontal="center", vertical="center")
        # Color code priority cell
        p_val = task.priority.upper()
        if p_val == "P1":
            c2.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif p_val == "P2":
            c2.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        elif p_val == "P3":
            c2.font = Font(name=font_family, size=11, bold=True, color="D6A23A")
        else:
            c2.font = Font(name=font_family, size=11, bold=True, color="3B82F6")
            
        c3 = ws3.cell(row=r, column=3, value=task.maintenance_category.capitalize() if task.maintenance_category else "Unclassified")
        c3.alignment = Alignment(horizontal="left", vertical="center")
        
        c4 = ws3.cell(row=r, column=4, value=task.recommendation)
        c4.alignment = Alignment(horizontal="left", vertical="center")
        
        c5 = ws3.cell(row=r, column=5, value=task.estimated_cost or 0)
        c5.number_format = "₹#,##0"
        c5.alignment = Alignment(horizontal="right", vertical="center")
        
        c6 = ws3.cell(row=r, column=6, value=task.estimated_response_time or "N/A")
        c6.alignment = Alignment(horizontal="left", vertical="center")
        
        c7 = ws3.cell(row=r, column=7, value=task.status.capitalize())
        c7.alignment = Alignment(horizontal="center", vertical="center")
        
        c8 = ws3.cell(row=r, column=8, value=due_str)
        c8.alignment = Alignment(horizontal="center", vertical="center")
        
        c9 = ws3.cell(row=r, column=9, value=engineer_name)
        c9.alignment = Alignment(horizontal="left", vertical="center")

        # Zebra striping & borders
        row_fill = PatternFill(start_color=color_zebra if row_idx % 2 == 1 else "FFFFFF", end_color=color_zebra if row_idx % 2 == 1 else "FFFFFF", fill_type="solid")
        for col_idx in range(1, 10):
            cell = ws3.cell(row=r, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx != 2:
                cell.font = regular_font

    # Column dimensions for Sheet 3 (excluding banner rows in measurements)
    for col in ws3.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ----------------------------------------------------
    # SHEET 4: Chainage Analysis
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="Chainage Analysis")
    ws4.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Target: group detections into 100-meter chainage intervals dynamically
    speed_mps = 13.889  # 50 km/h vehicle speed
    
    # Pre-fetch maintenance tasks
    task_distress_ids = [d.id for d in distresses]
    tasks = db.query(MaintenanceTask).filter(MaintenanceTask.distress_id.in_(task_distress_ids)).all() if task_distress_ids else []
    task_map = {t.distress_id: t for t in tasks}

    # Calculate chainage distance for each distress log
    distress_chainages = []
    for d in distresses:
        ts = d.video_timestamp if d.video_timestamp is not None else 0.0
        dist = ts * speed_mps
        distress_chainages.append((d, dist))

    # Calculate maximum distance to size the intervals
    max_dist = max([dist for _, dist in distress_chainages] + [10.0])
    interval_length = 100.0  # meters
    num_intervals = int(max_dist / interval_length) + 1

    # Populate intervals
    intervals_data = []
    for i in range(num_intervals):
        start_m = i * interval_length
        end_m = (i + 1) * interval_length
        
        # Filter distresses in this interval range
        interval_items = [d for d, dist in distress_chainages if start_m <= dist < end_m]
        
        # We only display intervals that contain distresses to maintain high clean readability
        if not interval_items:
            continue
            
        count = len(interval_items)
        crit_c = sum(1 for d in interval_items if d.severity.lower() == "critical")
        high_c = sum(1 for d in interval_items if d.severity.lower() == "high")
        med_c = sum(1 for d in interval_items if d.severity.lower() == "medium")
        low_c = sum(1 for d in interval_items if d.severity.lower() == "low")
        
        # Road Health Score for this interval
        health = max(0, 100 - (crit_c * 15 + high_c * 8 + med_c * 4 + low_c * 2))
        
        # Priority Rank
        if crit_c > 0:
            priority = "P1"
        elif high_c > 0:
            priority = "P2"
        elif med_c > 0:
            priority = "P3"
        else:
            priority = "P4"
            
        # Maintenance Recommendation (dominant recommendation)
        item_tasks = [task_map[d.id] for d in interval_items if d.id in task_map]
        recommendations = [t.recommendation for t in item_tasks if t.recommendation]
        if recommendations:
            # Mode recommendation
            rec = max(set(recommendations), key=recommendations.count)
        else:
            rec = "Monitor segment conditions."
            
        # Sum estimated rehabilitation cost
        cost = sum(t.estimated_cost for t in item_tasks if t.estimated_cost)
        
        # Chainage text labels
        start_str = f"KM {int(start_m // 1000)}+{int(start_m % 1000):03d}"
        end_str = f"KM {int(end_m // 1000)}+{int(end_m % 1000):03d}"
        
        intervals_data.append({
            "start": start_str,
            "end": end_str,
            "count": count,
            "critical": crit_c,
            "health": health,
            "priority": priority,
            "rec": rec,
            "cost": cost
        })

    # Title Banner (merged A1:H2)
    ws4.merge_cells("A1:H2")
    for r in range(1, 3):
        for c in range(1, 9):
            ws4.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws4.cell(row=r, column=c).border = thin_border
            
    title_cell = ws4.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - SPATIAL CHAINAGE ANALYSIS"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table Headers (Row 4)
    table_headers = [
        "Start Chainage", "End Chainage", "Number of Distresses", "Critical Count",
        "Road Health Index", "Priority Rank", "Maintenance Recommendation", "Estimated Cost"
    ]
    for col_idx, text in enumerate(table_headers):
        cell = ws4.cell(row=4, column=col_idx + 1, value=text)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws4.row_dimensions[4].height = 26

    # Populate table data starting at Row 5
    start_row = 5
    for idx, s in enumerate(intervals_data):
        r = start_row + idx
        ws4.row_dimensions[r].height = 20
        
        c1 = ws4.cell(row=r, column=1, value=s["start"])
        c1.font = bold_font
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        c2 = ws4.cell(row=r, column=2, value=s["end"])
        c2.font = bold_font
        c2.alignment = Alignment(horizontal="center", vertical="center")
        
        c3 = ws4.cell(row=r, column=3, value=s["count"])
        c3.font = regular_font
        c3.alignment = Alignment(horizontal="right", vertical="center")
        
        c4 = ws4.cell(row=r, column=4, value=s["critical"])
        c4.font = regular_font
        c4.alignment = Alignment(horizontal="right", vertical="center")
        
        # Road Health Index with Custom Conditional Styling Fills (green/orange/red)
        c5 = ws4.cell(row=r, column=5, value=s["health"])
        c5.alignment = Alignment(horizontal="right", vertical="center")
        if s["health"] >= 80:
            c5.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") # Light green
            c5.font = Font(name=font_family, size=11, bold=True, color="2E7D32")
        elif s["health"] >= 50:
            c5.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid") # Light orange
            c5.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        else:
            c5.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid") # Light red
            c5.font = Font(name=font_family, size=11, bold=True, color="C62828")
            
        c6 = ws4.cell(row=r, column=6, value=s["priority"])
        c6.alignment = Alignment(horizontal="center", vertical="center")
        # Color coding priority font
        p_val = s["priority"].upper()
        if p_val == "P1":
            c6.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif p_val == "P2":
            c6.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        elif p_val == "P3":
            c6.font = Font(name=font_family, size=11, bold=True, color="D6A23A")
        else:
            c6.font = Font(name=font_family, size=11, bold=True, color="3B82F6")
            
        c7 = ws4.cell(row=r, column=7, value=s["rec"])
        c7.font = regular_font
        c7.alignment = Alignment(horizontal="left", vertical="center")
        
        c8 = ws4.cell(row=r, column=8, value=s["cost"])
        c8.font = regular_font
        c8.number_format = "₹#,##0"
        c8.alignment = Alignment(horizontal="right", vertical="center")

        # Zebra backgrounds for other cells
        fill_color = color_zebra if idx % 2 == 1 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx in range(1, 9):
            cell = ws4.cell(row=r, column=col_idx)
            # Do not overwrite the conditional formatting fill of Column 5 (E)
            if col_idx != 5:
                cell.fill = row_fill
            cell.border = thin_border

    # Totals bottom row
    total_row = start_row + len(intervals_data)
    ws4.row_dimensions[total_row].height = 22
    
    t1 = ws4.cell(row=total_row, column=1, value="Total Summary")
    t1.font = bold_font
    t1.alignment = Alignment(horizontal="left", vertical="center")
    
    ws4.cell(row=total_row, column=2, value="")
    
    t3 = ws4.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row - 1})")
    t3.font = bold_font
    t3.alignment = Alignment(horizontal="right", vertical="center")
    
    t4 = ws4.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row - 1})")
    t4.font = bold_font
    t4.alignment = Alignment(horizontal="right", vertical="center")
    
    t5 = ws4.cell(row=total_row, column=5, value=f"=AVERAGE(E5:E{total_row - 1})")
    t5.font = bold_font
    t5.number_format = "0"
    t5.alignment = Alignment(horizontal="right", vertical="center")
    
    ws4.cell(row=total_row, column=6, value="")
    ws4.cell(row=total_row, column=7, value="")
    
    t8 = ws4.cell(row=total_row, column=8, value=f"=SUM(H5:H{total_row - 1})")
    t8.font = bold_font
    t8.number_format = "₹#,##0"
    t8.alignment = Alignment(horizontal="right", vertical="center")

    total_fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
    for col_idx in range(1, 9):
        cell = ws4.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = double_bottom_border

    # Auto-adjust column widths for Sheet 4 (excluding title banner)
    for col in ws4.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws4.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ----------------------------------------------------
    # SHEET 5: Cost Estimation
    # ----------------------------------------------------
    ws5 = wb.create_sheet(title="Cost Estimation")
    ws5.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Calculate last rows of data sheets for formula ranges
    last_row_plan = max(5, 4 + tasks_populated)
    last_row_details = max(5, 4 + len(distresses))

    # Dynamic lookup categories
    distress_types = sorted(list(set(d.distress_type.capitalize() for d in distresses))) if distresses else ["Pothole", "Alligator Cracks", "Ruts", "Corrugation"]
    severities = ["Critical", "High", "Medium", "Low"]
    categories = sorted(list(set(t.maintenance_category.capitalize() for t in tasks if t.maintenance_category))) if tasks else ["Crack sealing", "Patching", "Surfacing overlay", "Preventive sealing"]

    # Title Banner (merged A1:I2)
    ws5.merge_cells("A1:I2")
    for r in range(1, 3):
        for c in range(1, 10):
            ws5.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws5.cell(row=r, column=c).border = thin_border
            
    title_cell = ws5.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - MAINTENANCE COST ESTIMATION SUMMARY"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI Summary Cards (Row 4 & 5)
    # Card 1: Total Budget (Col A-B)
    ws5.merge_cells("A4:B4")
    lbl_1 = ws5.cell(row=4, column=1, value="Total Maintenance Budget")
    lbl_1.font = Font(name=font_family, size=9, bold=True, color="595959")
    lbl_1.alignment = Alignment(horizontal="center", vertical="center")
    
    ws5.merge_cells("A5:B5")
    val_1 = ws5.cell(row=5, column=1, value=f"=SUM('Maintenance Plan'!E5:E{last_row_plan})")
    val_1.font = Font(name=font_family, size=13, bold=True, color=color_primary)
    val_1.number_format = "₹#,##0"
    val_1.alignment = Alignment(horizontal="center", vertical="center")

    # Card 2: Average Task Cost (Col D-E)
    ws5.merge_cells("D4:E4")
    lbl_2 = ws5.cell(row=4, column=4, value="Average Repair Task Cost")
    lbl_2.font = Font(name=font_family, size=9, bold=True, color="595959")
    lbl_2.alignment = Alignment(horizontal="center", vertical="center")
    
    ws5.merge_cells("D5:E5")
    val_2 = ws5.cell(row=5, column=4, value=f"=AVERAGE('Maintenance Plan'!E5:E{last_row_plan})")
    val_2.font = Font(name=font_family, size=13, bold=True, color=color_primary)
    val_2.number_format = "₹#,##0"
    val_2.alignment = Alignment(horizontal="center", vertical="center")

    # Card 3: Maximum Task Cost (Col G-H)
    ws5.merge_cells("G4:H4")
    lbl_3 = ws5.cell(row=4, column=7, value="Maximum Single Repair Cost")
    lbl_3.font = Font(name=font_family, size=9, bold=True, color="595959")
    lbl_3.alignment = Alignment(horizontal="center", vertical="center")
    
    ws5.merge_cells("G5:H5")
    val_3 = ws5.cell(row=5, column=7, value=f"=MAX('Maintenance Plan'!E5:E{last_row_plan})")
    val_3.font = Font(name=font_family, size=13, bold=True, color="C62828")
    val_3.number_format = "₹#,##0"
    val_3.alignment = Alignment(horizontal="center", vertical="center")

    # Style cards block
    card_fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
    for r in [4, 5]:
        for c in range(1, 10):
            cell = ws5.cell(row=r, column=c)
            # Give background fill and borders to the columns containing cards
            if c in [1, 2, 4, 5, 7, 8]:
                cell.fill = card_fill
                cell.border = thin_border

    # Table section headers (Row 7)
    ws5.cell(row=7, column=1, value="Cost by Distress Category").font = section_font
    ws5.cell(row=7, column=4, value="Cost by Severity Rank").font = section_font
    ws5.cell(row=7, column=7, value="Cost by Repair Action Type").font = section_font

    # Column Headers (Row 8)
    # Table 1: Distress Type
    ws5.cell(row=8, column=1, value="Distress Type").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=1).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=2, value="Cost").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=2).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=2).alignment = Alignment(horizontal="right")

    # Table 2: Severity
    ws5.cell(row=8, column=4, value="Severity").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=4).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=5, value="Cost").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=5).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=5).alignment = Alignment(horizontal="right")

    # Table 3: Maintenance Category
    ws5.cell(row=8, column=7, value="Repair Category").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=7).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=8, value="Cost").font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    ws5.cell(row=8, column=8).fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    ws5.cell(row=8, column=8).alignment = Alignment(horizontal="right")

    for c in [1, 2, 4, 5, 7, 8]:
        ws5.cell(row=8, column=c).border = thin_border

    # Populate Data side-by-side starting at Row 9
    max_len = max(len(distress_types), len(severities), len(categories))
    for idx in range(max_len):
        r = 9 + idx
        ws5.row_dimensions[r].height = 20
        fill_color = color_zebra if idx % 2 == 1 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # 1. Distress Type Table
        if idx < len(distress_types):
            dtype = distress_types[idx]
            cell_lbl = ws5.cell(row=r, column=1, value=dtype)
            cell_lbl.font = regular_font
            cell_lbl.fill = row_fill
            cell_lbl.border = thin_border
            
            cell_val = ws5.cell(row=r, column=2, value=f"=SUMIFS('Detection Details'!N$5:N${last_row_details}, 'Detection Details'!C$5:C${last_row_details}, \"{dtype}\")")
            cell_val.font = regular_font
            cell_val.number_format = "₹#,##0"
            cell_val.alignment = Alignment(horizontal="right", vertical="center")
            cell_val.fill = row_fill
            cell_val.border = thin_border
            
        # 2. Severity Level Table
        if idx < len(severities):
            sev = severities[idx]
            cell_lbl = ws5.cell(row=r, column=4, value=sev)
            cell_lbl.font = regular_font
            cell_lbl.fill = row_fill
            cell_lbl.border = thin_border
            
            cell_val = ws5.cell(row=r, column=5, value=f"=SUMIFS('Detection Details'!N$5:N${last_row_details}, 'Detection Details'!D$5:D${last_row_details}, \"{sev}\")")
            cell_val.font = regular_font
            cell_val.number_format = "₹#,##0"
            cell_val.alignment = Alignment(horizontal="right", vertical="center")
            cell_val.fill = row_fill
            cell_val.border = thin_border
            
        # 3. Repair Action Category Table
        if idx < len(categories):
            cat = categories[idx]
            cell_lbl = ws5.cell(row=r, column=7, value=cat)
            cell_lbl.font = regular_font
            cell_lbl.fill = row_fill
            cell_lbl.border = thin_border
            
            cell_val = ws5.cell(row=r, column=8, value=f"=SUMIFS('Maintenance Plan'!E$5:E${last_row_plan}, 'Maintenance Plan'!C$5:C${last_row_plan}, \"{cat}\")")
            cell_val.font = regular_font
            cell_val.number_format = "₹#,##0"
            cell_val.alignment = Alignment(horizontal="right", vertical="center")
            cell_val.fill = row_fill
            cell_val.border = thin_border

    # Totals Row at the end
    tot_row = 9 + max_len
    ws5.row_dimensions[tot_row].height = 22
    total_fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
    
    # Distress type total
    t1_lbl = ws5.cell(row=tot_row, column=1, value="Total")
    t1_lbl.font = bold_font
    t1_lbl.fill = total_fill
    t1_lbl.border = double_bottom_border
    
    t1_val = ws5.cell(row=tot_row, column=2, value=f"=SUM(B9:B{tot_row - 1})")
    t1_val.font = bold_font
    t1_val.number_format = "₹#,##0"
    t1_val.alignment = Alignment(horizontal="right", vertical="center")
    t1_val.fill = total_fill
    t1_val.border = double_bottom_border

    # Severity total
    t2_lbl = ws5.cell(row=tot_row, column=4, value="Total")
    t2_lbl.font = bold_font
    t2_lbl.fill = total_fill
    t2_lbl.border = double_bottom_border
    
    t2_val = ws5.cell(row=tot_row, column=5, value=f"=SUM(E9:E{tot_row - 1})")
    t2_val.font = bold_font
    t2_val.number_format = "₹#,##0"
    t2_val.alignment = Alignment(horizontal="right", vertical="center")
    t2_val.fill = total_fill
    t2_val.border = double_bottom_border

    # Category total
    t3_lbl = ws5.cell(row=tot_row, column=7, value="Total")
    t3_lbl.font = bold_font
    t3_lbl.fill = total_fill
    t3_lbl.border = double_bottom_border
    
    t3_val = ws5.cell(row=tot_row, column=8, value=f"=SUM(H9:H{tot_row - 1})")
    t3_val.font = bold_font
    t3_val.number_format = "₹#,##0"
    t3_val.alignment = Alignment(horizontal="right", vertical="center")
    t3_val.fill = total_fill
    t3_val.border = double_bottom_border

    # Set column widths for Sheet 5 (excluding title banner)
    for col in ws5.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 4, 5] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws5.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ----------------------------------------------------
    # SHEET 6: GIS Data
    # ----------------------------------------------------
    ws6 = wb.create_sheet(title="GIS")
    ws6.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Location naming hashes
    routes_list = ["NH-48", "NH-4", "SH-10", "NH-66"]
    roads_list = ["Mumbai-Pune Expressway", "Western Express Highway", "Sion-Panvel Highway", "Eastern Express Highway"]
    road_name = roads_list[video_id % len(roads_list)]
    route_name = routes_list[video_id % len(roads_list)]
    
    speed_mps = 13.889 # 50 km/h vehicle speed

    # Merged Banner Header (A1:J2)
    ws6.merge_cells("A1:J2")
    for r in range(1, 3):
        for c in range(1, 11):
            ws6.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws6.cell(row=r, column=c).border = thin_border
            
    title_cell = ws6.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - GEOGRAPHIC INFORMATION SYSTEM (GIS) DATA"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    gis_headers = [
        "Tracking ID", "Latitude", "Longitude", "Chainage", "Road Name", 
        "Distress Type", "Severity", "Priority", "Status", "Video Timestamp"
    ]
    
    # Headers formatting at Row 4
    for col_idx, text in enumerate(gis_headers):
        cell = ws6.cell(row=4, column=col_idx + 1, value=text)
        cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws6.row_dimensions[4].height = 26

    for idx, d in enumerate(distresses):
        r = 5 + idx
        ws6.row_dimensions[r].height = 20
        task = task_map.get(d.id)
        
        # Calculate chainage distance meters
        ts = d.video_timestamp if d.video_timestamp is not None else 0.0
        dist = ts * speed_mps
        chainage_str = f"KM {int(dist // 1000)}+{int(dist % 1000):03d}"
        
        # Resolve priority & status values
        priority_val = task.priority.upper() if task else ("P1" if d.severity.lower() == "critical" else ("P2" if d.severity.lower() == "high" else ("P3" if d.severity.lower() == "medium" else "P4")))
        status_val = task.status.capitalize() if task else d.status.capitalize()
        
        # Severity color font styling
        sev_lower = d.severity.lower()
        if sev_lower == "critical":
            sev_color = "C62828"
        elif sev_lower == "high":
            sev_color = "EF6C00"
        elif sev_lower == "medium":
            sev_color = "D6A23A"
        else:
            sev_color = "3B82F6"

        # Populate columns
        c1 = ws6.cell(row=r, column=1, value=d.tracking_id if d.tracking_id is not None else "N/A")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        
        c2 = ws6.cell(row=r, column=2, value=d.latitude)
        c2.number_format = "0.00000"
        c2.alignment = Alignment(horizontal="right", vertical="center")
        
        c3 = ws6.cell(row=r, column=3, value=d.longitude)
        c3.number_format = "0.00000"
        c3.alignment = Alignment(horizontal="right", vertical="center")
        
        c4 = ws6.cell(row=r, column=4, value=chainage_str)
        c4.alignment = Alignment(horizontal="center", vertical="center")
        
        c5 = ws6.cell(row=r, column=5, value=road_name)
        c5.alignment = Alignment(horizontal="left", vertical="center")
        
        c6 = ws6.cell(row=r, column=6, value=d.distress_type.capitalize())
        c6.alignment = Alignment(horizontal="left", vertical="center")
        
        c7 = ws6.cell(row=r, column=7, value=d.severity.capitalize())
        c7.font = Font(name=font_family, size=11, bold=True, color=sev_color)
        c7.alignment = Alignment(horizontal="center", vertical="center")
        
        c8 = ws6.cell(row=r, column=8, value=priority_val)
        c8.alignment = Alignment(horizontal="center", vertical="center")
        # Color coding priority rank
        if priority_val == "P1":
            c8.font = Font(name=font_family, size=11, bold=True, color="C62828")
        elif priority_val == "P2":
            c8.font = Font(name=font_family, size=11, bold=True, color="EF6C00")
        elif priority_val == "P3":
            c8.font = Font(name=font_family, size=11, bold=True, color="D6A23A")
        else:
            c8.font = Font(name=font_family, size=11, bold=True, color="3B82F6")
            
        c9 = ws6.cell(row=r, column=9, value=status_val)
        c9.alignment = Alignment(horizontal="center", vertical="center")
        
        c10 = ws6.cell(row=r, column=10, value=d.video_timestamp)
        if d.video_timestamp is not None:
            c10.number_format = "0.0"
        c10.alignment = Alignment(horizontal="right", vertical="center")

        # Zebra striping & borders
        row_fill = PatternFill(start_color=color_zebra if idx % 2 == 1 else "FFFFFF", end_color=color_zebra if idx % 2 == 1 else "FFFFFF", fill_type="solid")
        for col_idx in range(1, 11):
            cell = ws6.cell(row=r, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx not in [7, 8]:
                cell.font = regular_font

    # Set column widths for Sheet 6 (excluding title banner)
    for col in ws6.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2] or cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws6.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # ----------------------------------------------------
    # SHEET 7: Charts
    # ----------------------------------------------------
    ws7 = wb.create_sheet(title="Charts")
    ws7.views.sheetView[0].showGridLines = True

    # Theme colors definition
    color_primary = "3E4331"  # Dark Olive
    color_header = "6C7354"   # Medium Olive
    color_accent = "F1EEE5"   # Warm Ivory
    color_zebra = "FAF8F2"    # Light Zebra

    # Dynamic metrics calculations
    p1_count = sum(1 for t in tasks if t.priority.upper() == "P1")
    p2_count = sum(1 for t in tasks if t.priority.upper() == "P2")
    p3_count = sum(1 for t in tasks if t.priority.upper() == "P3")
    p4_count = sum(1 for t in tasks if t.priority.upper() == "P4")

    conf_90_100 = sum(1 for d in distresses if d.confidence_score >= 0.9)
    conf_75_89 = sum(1 for d in distresses if 0.75 <= d.confidence_score < 0.9)
    conf_60_74 = sum(1 for d in distresses if 0.6 <= d.confidence_score < 0.75)
    conf_lt_60 = sum(1 for d in distresses if d.confidence_score < 0.6)

    completed_videos = db.query(UploadedVideo).filter(UploadedVideo.processing_status == "completed").all()
    avg_proc_time = sum(v.processing_duration for v in completed_videos if v.processing_duration) / len(completed_videos) if completed_videos else 0.0

    cat_costs = {}
    for cat in categories:
        cat_costs[cat] = sum(t.estimated_cost for t in tasks if t.estimated_cost and t.maintenance_category and t.maintenance_category.lower() == cat.lower())

    distress_types_dist = {}
    for d in distresses:
        t_name = d.distress_type.capitalize()
        distress_types_dist[t_name] = distress_types_dist.get(t_name, 0) + 1

    # Title Banner (merged A1:I2)
    ws7.merge_cells("A1:I2")
    for r in range(1, 3):
        for c in range(1, 10):
            ws7.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
            ws7.cell(row=r, column=c).border = thin_border
            
    title_cell = ws7.cell(row=1, column=1)
    title_cell.value = "ROAD DISTRESS AUDIT - SYSTEM CHART REPORT"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Stack data tables in Columns A-B dynamically to feed the charts
    tb_fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")

    # Table 1: Distress Category Distribution
    r_distress_start = 3
    ws7.cell(row=r_distress_start, column=1, value="Distress Category").font = bold_font
    ws7.cell(row=r_distress_start, column=1).fill = tb_fill
    ws7.cell(row=r_distress_start, column=2, value="Anomalies").font = bold_font
    ws7.cell(row=r_distress_start, column=2).fill = tb_fill
    
    dist_items = sorted(list(distress_types_dist.items()))
    for idx, (dtype, count) in enumerate(dist_items):
        row_num = r_distress_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=dtype).font = regular_font
        ws7.cell(row=row_num, column=2, value=count).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border
    last_distress_row = r_distress_start + len(dist_items)

    # Table 2: Severity Distribution
    r_severity_start = last_distress_row + 2
    ws7.cell(row=r_severity_start, column=1, value="Severity Level").font = bold_font
    ws7.cell(row=r_severity_start, column=1).fill = tb_fill
    ws7.cell(row=r_severity_start, column=2, value="Count").font = bold_font
    ws7.cell(row=r_severity_start, column=2).fill = tb_fill
    
    sev_items = [("Critical", critical_count), ("High", high_count), ("Medium", medium_count), ("Low", low_count)]
    for idx, (sev, count) in enumerate(sev_items):
        row_num = r_severity_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=sev).font = regular_font
        ws7.cell(row=row_num, column=2, value=count).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border
    last_severity_row = r_severity_start + 4

    # Table 3: Priority Distribution
    r_priority_start = last_severity_row + 2
    ws7.cell(row=r_priority_start, column=1, value="Priority Rank").font = bold_font
    ws7.cell(row=r_priority_start, column=1).fill = tb_fill
    ws7.cell(row=r_priority_start, column=2, value="Count").font = bold_font
    ws7.cell(row=r_priority_start, column=2).fill = tb_fill
    
    prio_items = [("P1", p1_count), ("P2", p2_count), ("P3", p3_count), ("P4", p4_count)]
    for idx, (prio, count) in enumerate(prio_items):
        row_num = r_priority_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=prio).font = regular_font
        ws7.cell(row=row_num, column=2, value=count).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border
    last_priority_row = r_priority_start + 4

    # Table 4: Road Health Gauge Rating
    r_health_start = last_priority_row + 2
    ws7.cell(row=r_health_start, column=1, value="Metric").font = bold_font
    ws7.cell(row=r_health_start, column=1).fill = tb_fill
    ws7.cell(row=r_health_start, column=2, value="Score").font = bold_font
    ws7.cell(row=r_health_start, column=2).fill = tb_fill
    
    ws7.cell(row=r_health_start + 1, column=1, value="Road Health Index").font = regular_font
    ws7.cell(row=r_health_start + 1, column=2, value=road_health_score).font = regular_font
    ws7.cell(row=r_health_start + 1, column=1).border = thin_border
    ws7.cell(row=r_health_start + 1, column=2).border = thin_border
    last_health_row = r_health_start + 1

    # Table 5: Maintenance Category Costs
    r_cost_start = last_health_row + 2
    ws7.cell(row=r_cost_start, column=1, value="Repair Category").font = bold_font
    ws7.cell(row=r_cost_start, column=1).fill = tb_fill
    ws7.cell(row=r_cost_start, column=2, value="Total Cost").font = bold_font
    ws7.cell(row=r_cost_start, column=2).fill = tb_fill
    
    cost_items = sorted(list(cat_costs.items()))[:4] # limit to top 4 categories
    for idx, (cat, cost) in enumerate(cost_items):
        row_num = r_cost_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=cat).font = regular_font
        ws7.cell(row=row_num, column=2, value=cost).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border
    last_cost_row = r_cost_start + len(cost_items)

    # Table 6: Confidence Score Distribution
    r_conf_start = last_cost_row + 2
    ws7.cell(row=r_conf_start, column=1, value="Confidence range").font = bold_font
    ws7.cell(row=r_conf_start, column=1).fill = tb_fill
    ws7.cell(row=r_conf_start, column=2, value="Detections").font = bold_font
    ws7.cell(row=r_conf_start, column=2).fill = tb_fill
    
    conf_items = [("90% - 100%", conf_90_100), ("75% - 89%", conf_75_89), ("60% - 74%", conf_60_74), ("< 60%", conf_lt_60)]
    for idx, (rng, count) in enumerate(conf_items):
        row_num = r_conf_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=rng).font = regular_font
        ws7.cell(row=row_num, column=2, value=count).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border
    last_conf_row = r_conf_start + 4

    # Table 7: Processing Statistics
    r_stats_start = last_conf_row + 2
    ws7.cell(row=r_stats_start, column=1, value="Processing Metric").font = bold_font
    ws7.cell(row=r_stats_start, column=1).fill = tb_fill
    ws7.cell(row=r_stats_start, column=2, value="Time / Runs").font = bold_font
    ws7.cell(row=r_stats_start, column=2).fill = tb_fill
    
    stats_items = [
        ("Total Inspection Runs", total_videos_processed), 
        ("Video Analysis Time (s)", video.processing_duration or 0.0), 
        ("Average System Runtime (s)", avg_proc_time)
    ]
    for idx, (metric, val) in enumerate(stats_items):
        row_num = r_stats_start + 1 + idx
        ws7.cell(row=row_num, column=1, value=metric).font = regular_font
        ws7.cell(row=row_num, column=2, value=val).font = regular_font
        ws7.cell(row=row_num, column=1).border = thin_border
        ws7.cell(row=row_num, column=2).border = thin_border

    # Create & Position Native Excel Chart Elements dynamically
    # Chart 1: Distress Category Pie Chart (D4)
    if dist_items:
        pie = PieChart()
        pie.title = "Distress Category Distribution"
        pie.width = 11
        pie.height = 7
        data_ref = Reference(ws7, min_col=2, min_row=r_distress_start, max_row=last_distress_row)
        cats_ref = Reference(ws7, min_col=1, min_row=r_distress_start + 1, max_row=last_distress_row)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws7.add_chart(pie, "D4")

    # Chart 2: Severity Bar Chart (H4)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Anomalies by Severity Rank"
    chart2.y_axis.title = "Detections Count"
    chart2.x_axis.title = "Severity Level"
    chart2.width = 11
    chart2.height = 7
    chart2.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_severity_start, max_row=last_severity_row)
    cats_ref = Reference(ws7, min_col=1, min_row=r_severity_start + 1, max_row=last_severity_row)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    ws7.add_chart(chart2, "H4")

    # Chart 3: Priority Bar Chart (D18)
    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "Maintenance Tasks by Priority Rank"
    chart3.y_axis.title = "Schedules Count"
    chart3.x_axis.title = "Priority Rank"
    chart3.width = 11
    chart3.height = 7
    chart3.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_priority_start, max_row=last_priority_row)
    cats_ref = Reference(ws7, min_col=1, min_row=r_priority_start + 1, max_row=last_priority_row)
    chart3.add_data(data_ref, titles_from_data=True)
    chart3.set_categories(cats_ref)
    ws7.add_chart(chart3, "D18")

    # Chart 4: Road Health Gauge Index equivalent (H18)
    chart4 = BarChart()
    chart4.type = "col"
    chart4.title = "Road Health Index Rating"
    chart4.y_axis.title = "Index (0-100)"
    chart4.y_axis.scaling.max = 100
    chart4.y_axis.scaling.min = 0
    chart4.width = 11
    chart4.height = 7
    chart4.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_health_start, max_row=last_health_row)
    cats_ref = Reference(ws7, min_col=1, min_row=r_health_start + 1, max_row=last_health_row)
    chart4.add_data(data_ref, titles_from_data=True)
    chart4.set_categories(cats_ref)
    ws7.add_chart(chart4, "H18")

    # Chart 5: Category Cost Bar Chart (D32)
    chart5 = BarChart()
    chart5.type = "col"
    chart5.title = "Rehabilitation Cost by Action Category"
    chart5.y_axis.title = "Estimated Cost (₹)"
    chart5.x_axis.title = "Repair Action"
    chart5.width = 11
    chart5.height = 7
    chart5.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_cost_start, max_row=last_cost_row)
    cats_ref = Reference(ws7, min_col=1, min_row=r_cost_start + 1, max_row=last_cost_row)
    chart5.add_data(data_ref, titles_from_data=True)
    chart5.set_categories(cats_ref)
    ws7.add_chart(chart5, "D32")

    # Chart 6: AI Confidence Distribution Chart (H32)
    chart6 = BarChart()
    chart6.type = "col"
    chart6.title = "AI Confidence Score Distribution"
    chart6.y_axis.title = "Detections Count"
    chart6.x_axis.title = "Confidence Range"
    chart6.width = 11
    chart6.height = 7
    chart6.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_conf_start, max_row=last_conf_row)
    cats_ref = Reference(ws7, min_col=1, min_row=r_conf_start + 1, max_row=last_conf_row)
    chart6.add_data(data_ref, titles_from_data=True)
    chart6.set_categories(cats_ref)
    ws7.add_chart(chart6, "H32")

    # Chart 7: Processing Statistics Chart (D46)
    chart7 = BarChart()
    chart7.type = "col"
    chart7.title = "Video Analysis Runtime Performance"
    chart7.y_axis.title = "Value"
    chart7.x_axis.title = "Performance Indicators"
    chart7.width = 11
    chart7.height = 7
    chart7.legend = None
    data_ref = Reference(ws7, min_col=2, min_row=r_stats_start, max_row=r_stats_start + 3)
    cats_ref = Reference(ws7, min_col=1, min_row=r_stats_start + 1, max_row=r_stats_start + 3)
    chart7.add_data(data_ref, titles_from_data=True)
    chart7.set_categories(cats_ref)
    ws7.add_chart(chart7, "D46")

    # Width adjustment for Columns A and B reference tables
    ws7.column_dimensions["A"].width = 28
    ws7.column_dimensions["B"].width = 15

    # ----------------------------------------------------
    # SHEETS 8-11: NSV-style survey report tabs
    # ----------------------------------------------------
    # Modeled on a real Network Survey Vehicle (NSV) condition report format
    # (tab names/columns like "Summary", "100m", "Potholes", and the IRC:82
    # permissible-limit criteria table) supplied as a reference. That report
    # is built from laser profilometer/rut-bar sensor data on a highway
    # corridor; this system only has point-in-time AI video detections
    # (type, GPS, timestamp, confidence, bounding box). Wherever a real
    # figure exists in our own data it's used directly (counts, chainage
    # derived from GPS/timestamp, severity, damage area); a handful of
    # fields the video pipeline has no way to measure (pothole depth in mm,
    # lane number, highway number) are explicitly filled with a labeled
    # placeholder rather than a real reading -- flagged in orange with a
    # cell comment, never presented as a real measurement.
    color_primary = "1F4E78"
    color_header = "2E6DA4"
    color_accent = "D9E1F2"
    color_zebra = "F2F6FC"
    placeholder_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    placeholder_font = Font(name=font_family, size=11, italic=True, color="9C5700")

    from openpyxl.comments import Comment

    def placeholder_cell(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = placeholder_fill
        cell.font = placeholder_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(
            "Placeholder: not measurable from AI video detections (requires NSV/laser "
            "profilometer sensor survey equipment). Not a real reading.",
            "Road Distress Management System"
        )
        return cell

    REAL_CLASS_LABELS = {
        "longitudinal_crack": "Longitudinal Crack",
        "transverse_crack": "Transverse Crack",
        "alligator_crack": "Alligator Crack",
        "pothole": "Pothole",
        "poles": "Poles",
        "traffic sign": "Traffic Sign",
        "sign board": "Sign Board",
    }

    def banner(ws, text, last_col):
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col)
        for r in range(1, 3):
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
                ws.cell(row=r, column=c).border = thin_border
        cell = ws.cell(row=1, column=1, value=text)
        cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def header_row(ws, row, headers):
        for col_idx, text in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1, value=text)
            cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row].height = 26

    def autosize(ws, skip_rows):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in skip_rows or cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # --- Sheet 8: NSV Style - Permissible Limits Reference ---
    ws8 = wb.create_sheet(title="NSV Style - Criteria")
    banner(ws8, "PERMISSIBLE DISTRESS LIMITS - REFERENCE (IRC:82-2023)", 2)
    ws8.cell(row=4, column=1, value=(
        "Reproduced from the reference NSV survey report format's 'Repair and Maintenance asper CA' "
        "tab. These are the National Highway maintenance criteria under IRC:82-2023 -- a published "
        "standard, not project-specific data. This AI video pipeline detects individual distress "
        "instances (with GPS + timestamp), not the lane-length percentage these limits are defined "
        "against, so they're included here as reference only; see the 'NSV Style - Summary' tab for "
        "how this system's own real counts compare."
    )).font = italic_font
    ws8.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws8.merge_cells("A4:B4")
    ws8.row_dimensions[4].height = 60

    header_row(ws8, 6, ["S. No.", "Nature of Defect or Deficiency (Carriageway and Paved Shoulder)"])
    criteria_rows = [
        (1, "Roughness value exceeding 2400 mm/km as per IRC:82-2023"),
        (2, "Potholes"),
        (3, "Cracking exceeding 10% of road surface area as per IRC:82-2023"),
        (4, "Rutting exceeding 10mm road surface area as per IRC:82-2023"),
        (5, "Ravelling exceeding 10% of road surface as per IRC:82-2023"),
    ]
    for idx, (num, text) in enumerate(criteria_rows):
        r = 7 + idx
        fill_color = color_zebra if idx % 2 == 1 else "FFFFFF"
        c1 = ws8.cell(row=r, column=1, value=num)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2 = ws8.cell(row=r, column=2, value=text)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for c in (c1, c2):
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            c.border = thin_border
            c.font = regular_font
    ws8.column_dimensions["A"].width = 10
    ws8.column_dimensions["B"].width = 90

    # --- Sheet 9: NSV Style - Summary ---
    ws9 = wb.create_sheet(title="NSV Style - Summary")
    banner(ws9, "ROAD DISTRESS AUDIT - NSV-STYLE DISTRESS SUMMARY", 5)

    ws9.cell(row=4, column=1, value="Total Distress Instances Detected:").font = bold_font
    ws9.cell(row=4, column=2, value=len(distresses)).font = bold_font
    ws9.cell(row=5, column=1, value="Video:").font = bold_font
    ws9.cell(row=5, column=2, value=video.filename).font = regular_font

    header_row(ws9, 7, ["Distress Type", "Permissible Limit (IRC:82-2023)", "Total Detected", "Requiring Rectification (High/Critical)", "% Requiring Rectification"])

    type_counts = {}
    for d in distresses:
        key = d.distress_type.lower()
        type_counts.setdefault(key, []).append(d)

    # Real published limits only exist for the 4 sensor-measured categories
    # (defined as % of lane length, not applicable to point counts); every
    # other real class this pipeline detects has no published count-based
    # limit, so it's marked N/A rather than invented.
    limit_map = {
        "pothole": "See IRC:82-2023 (no count-based limit published)",
    }

    row_cursor = 8
    for key in sorted(type_counts.keys()):
        items = type_counts[key]
        label = REAL_CLASS_LABELS.get(key, key.replace("_", " ").title())
        total = len(items)
        needs_fix = sum(1 for d in items if d.severity.lower() in ("high", "critical"))
        pct = needs_fix / total if total else 0.0

        c1 = ws9.cell(row=row_cursor, column=1, value=label)
        c1.font = bold_font
        c2 = ws9.cell(row=row_cursor, column=2, value=limit_map.get(key, "N/A (not a sensor-measured category)"))
        c2.font = italic_font
        c3 = ws9.cell(row=row_cursor, column=3, value=total)
        c4 = ws9.cell(row=row_cursor, column=4, value=needs_fix)
        c5 = ws9.cell(row=row_cursor, column=5, value=pct)
        c5.number_format = "0.0%"
        fill_color = color_zebra if row_cursor % 2 == 0 else "FFFFFF"
        for c in (c1, c2, c3, c4, c5):
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if c is c2 else Alignment(horizontal="left" if c is c1 else "center", vertical="center")
        row_cursor += 1

    total_row = row_cursor
    ws9.cell(row=total_row, column=1, value="All Distress Types").font = bold_font
    ws9.cell(row=total_row, column=3, value=f"=SUM(C8:C{total_row - 1})").font = bold_font
    ws9.cell(row=total_row, column=4, value=f"=SUM(D8:D{total_row - 1})").font = bold_font
    t5 = ws9.cell(row=total_row, column=5, value=f"=IFERROR(D{total_row}/C{total_row},0)")
    t5.font = bold_font
    t5.number_format = "0.0%"
    for c in range(1, 6):
        cell = ws9.cell(row=total_row, column=c)
        cell.fill = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
        cell.border = double_bottom_border
    autosize(ws9, {1, 2, 4, 5})
    ws9.column_dimensions["B"].width = 34

    # --- Sheet 10: NSV Style - 100m Segments ---
    ws10 = wb.create_sheet(title="NSV Style - 100m")
    class_keys_sorted = sorted(type_counts.keys())
    seg_headers = ["NH No.", "Start Chainage", "End Chainage", "Length (m)"] + \
        [REAL_CLASS_LABELS.get(k, k.replace("_", " ").title()) for k in class_keys_sorted] + \
        ["Dominant Severity", "Priority Rank"]
    banner(ws10, "ROAD DISTRESS AUDIT - NSV-STYLE 100m CHAINAGE SEGMENTS", len(seg_headers))
    ws10.cell(row=4, column=1, value=(
        "NH No. is a placeholder (this pipeline does not capture a highway route number); chainage is "
        "derived from each detection's real video timestamp at an assumed 50 km/h survey speed, the same "
        "method already used on the 'Chainage Analysis' tab."
    )).font = italic_font
    ws10.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(seg_headers))
    ws10.row_dimensions[4].height = 28

    header_row(ws10, 6, seg_headers)

    seg_row = 7
    for i in range(num_intervals):
        start_m = i * interval_length
        end_m = (i + 1) * interval_length
        interval_items = [d for d, dist in distress_chainages if start_m <= dist < end_m]
        if not interval_items:
            continue

        counts_by_class = {k: 0 for k in class_keys_sorted}
        for d in interval_items:
            key = d.distress_type.lower()
            if key in counts_by_class:
                counts_by_class[key] += 1

        severities = [d.severity.lower() for d in interval_items]
        if "critical" in severities:
            dom_sev, priority = "Critical", "P1"
        elif "high" in severities:
            dom_sev, priority = "High", "P2"
        elif "medium" in severities:
            dom_sev, priority = "Medium", "P3"
        else:
            dom_sev, priority = "Low", "P4"

        placeholder_cell(ws10, seg_row, 1, "NH0000")
        c2 = ws10.cell(row=seg_row, column=2, value=f"{int(start_m)}")
        c3 = ws10.cell(row=seg_row, column=3, value=f"{int(end_m)}")
        c4 = ws10.cell(row=seg_row, column=4, value=end_m - start_m)
        for c in (c2, c3, c4):
            c.alignment = Alignment(horizontal="center", vertical="center")

        col_idx = 5
        for k in class_keys_sorted:
            cell = ws10.cell(row=seg_row, column=col_idx, value=counts_by_class[k])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        sev_col = ws10.cell(row=seg_row, column=col_idx, value=dom_sev)
        sev_color = {"Critical": "C62828", "High": "EF6C00", "Medium": "D6A23A", "Low": "3B82F6"}[dom_sev]
        sev_col.font = Font(name=font_family, size=11, bold=True, color=sev_color)
        sev_col.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1
        prio_col = ws10.cell(row=seg_row, column=col_idx, value=priority)
        prio_col.font = Font(name=font_family, size=11, bold=True, color=sev_color)
        prio_col.alignment = Alignment(horizontal="center", vertical="center")

        fill_color = color_zebra if seg_row % 2 == 0 else "FFFFFF"
        for c in range(2, len(seg_headers) + 1):
            cell = ws10.cell(row=seg_row, column=c)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = thin_border
        seg_row += 1

    if seg_row == 7:
        ws10.cell(row=7, column=1, value="No geo-tagged detections available to segment.").font = italic_font
    autosize(ws10, {1, 4})
    ws10.freeze_panes = "A7"

    # --- Sheet 11: NSV Style - Potholes ---
    ws11 = wb.create_sheet(title="NSV Style - Potholes")
    pothole_headers = ["NH No.", "Chainage (m)", "Lane No.", "Area (Sq m)", "Max Depth (mm)", "Average Depth (mm)", "Severity"]
    banner(ws11, "ROAD DISTRESS AUDIT - NSV-STYLE POTHOLE LOG", len(pothole_headers))
    ws11.cell(row=4, column=1, value=(
        "NH No. and Lane No. are placeholders (not captured by this pipeline). Area is the real damage "
        "area computed from the AI bounding box. Max/Average Depth require a laser depth sensor this "
        "video-based pipeline does not have, so they're placeholders, not real measurements."
    )).font = italic_font
    ws11.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(pothole_headers))
    ws11.row_dimensions[4].height = 40
    header_row(ws11, 6, pothole_headers)

    potholes = [d for d in distresses if d.distress_type.lower() == "pothole"]
    pothole_chainage = {d.id: dist for d, dist in distress_chainages}
    p_row = 7
    for idx, d in enumerate(sorted(potholes, key=lambda x: pothole_chainage.get(x.id, 0))):
        placeholder_cell(ws11, p_row, 1, "NH0000")
        c2 = ws11.cell(row=p_row, column=2, value=round(pothole_chainage.get(d.id, 0.0), 1))
        c2.alignment = Alignment(horizontal="center", vertical="center")
        placeholder_cell(ws11, p_row, 3, "L1")
        c4 = ws11.cell(row=p_row, column=4, value=d.affected_area)
        c4.number_format = "0.000"
        c4.alignment = Alignment(horizontal="center", vertical="center")
        placeholder_cell(ws11, p_row, 5, "N/A")
        placeholder_cell(ws11, p_row, 6, "N/A")
        c7 = ws11.cell(row=p_row, column=7, value=d.severity.capitalize())
        sev_color = {"critical": "C62828", "high": "EF6C00", "medium": "D6A23A", "low": "3B82F6"}.get(d.severity.lower(), "3B82F6")
        c7.font = Font(name=font_family, size=11, bold=True, color=sev_color)
        c7.alignment = Alignment(horizontal="center", vertical="center")

        fill_color = color_zebra if idx % 2 == 1 else "FFFFFF"
        for c in (c2, c4, c7):
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            c.border = thin_border
        p_row += 1

    if p_row == 7:
        ws11.cell(row=7, column=1, value="No pothole detections in this video.").font = italic_font
    autosize(ws11, {1, 4})
    ws11.freeze_panes = "A7"

    # ----------------------------------------------------
    # GLOBAL FORMATTING & METADATA UPGRADES
    # ----------------------------------------------------
    # Freeze panes lock configurations
    ws1.freeze_panes = "A10"
    ws3.freeze_panes = "A5"
    ws4.freeze_panes = "A5"
    ws5.freeze_panes = "A6"
    ws6.freeze_panes = "A5"

    # Configure Workbook Metadata properties
    wb.properties.title = "Road Distress Audit Report"
    wb.properties.creator = "Road Distress Management System AI Pipeline"
    wb.properties.subject = "Automated Infrastructure Health Report"
    wb.properties.keywords = "Road, Distress, AI, YOLO, GIS, Maintenance, Cost, Executive"

    # Configure Print Setup, Headers, and Footers for all worksheets
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = "9"  # A4 Size
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Header configurations
        ws.oddHeader.center.text = "ROAD DISTRESS MANAGEMENT SYSTEM - AUDIT REPORT"
        ws.oddHeader.center.size = 9
        ws.oddHeader.center.font = font_family
        
        # Footer configurations
        ws.oddFooter.left.text = "Confidential - Infrastructure Inspection Services"
        ws.oddFooter.left.size = 8
        ws.oddFooter.left.font = font_family
        
        ws.oddFooter.center.text = "Generated on: &D"
        ws.oddFooter.center.size = 8
        ws.oddFooter.center.font = font_family
        
        ws.oddFooter.right.text = "Page &P of &N"
        ws.oddFooter.right.size = 8
        ws.oddFooter.right.font = font_family

    # Save to file
    wb.save(full_path)
    logger.info(f"Generated professional Excel report saved at: {full_path}")
    return relative_path
